"""
Exportación Avanzada a Excel - Sprint 2.3
Genera reportes ejecutivos con múltiples pestañas, gráficos y formato profesional
"""

from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import Response
from sqlmodel import Session, select
from app.models import Evaluation, ControlResponse, ControlDefinition, Norma, Client
from app.auth import get_current_user
from datetime import datetime
import io

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/excel/{evaluation_id}")
def export_to_excel(evaluation_id: str, request: Request):
    """Exportar evaluación completa a Excel con formato ejecutivo"""
    session_id = request.cookies.get("session_id")
    user = get_current_user(session_id)
    if not user:
        raise HTTPException(status_code=401, detail="No autenticado")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference, PieChart
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no instalado. Ejecutar: pip install openpyxl",
        )

    with Session(engine) as session:
        evaluation = session.get(Evaluation, evaluation_id)
        if not evaluation:
            raise HTTPException(status_code=404, detail="Evaluación no encontrada")

        # Validar permisos
        if user.role.value != "superadmin" and evaluation.client_id != user.client_id:
            raise HTTPException(status_code=403, detail="No autorizado")

        # Obtener datos
        norma = session.get(Norma, evaluation.norma_id) if evaluation.norma_id else None
        client = (
            session.get(Client, evaluation.client_id) if evaluation.client_id else None
        )

        responses = session.exec(
            select(ControlResponse).where(
                ControlResponse.evaluation_id == evaluation_id
            )
        ).all()

        # Crear libro de Excel
        wb = Workbook()
        wb.title = f"{evaluation.name}"

        # === HOJA 1: RESUMEN EJECUTIVO ===
        ws_summary = wb.active
        ws_summary.title = "Resumen Ejecutivo"

        # Estilos
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(
            start_color="1e40af", end_color="1e40af", fill_type="solid"
        )
        header_fill = PatternFill(
            start_color="f1f5f9", end_color="f1f5f9", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Título
        ws_summary.merge_cells("A1:D1")
        cell = ws_summary["A1"]
        cell.value = f"REPORTE DE EVALUACIÓN - {evaluation.name}"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Info básica
        ws_summary["A3"] = "Cliente:"
        ws_summary["B3"] = client.name if client else "N/A"
        ws_summary["A4"] = "Norma:"
        ws_summary["B4"] = norma.name if norma else "N/A"
        ws_summary["A5"] = "Fecha:"
        ws_summary["B5"] = evaluation.created_at.strftime("%d/%m/%Y")
        ws_summary["A6"] = "Total Controles:"
        ws_summary["B6"] = len(responses)

        # Calcular métricas
        total = len(responses)
        compliant = sum(1 for r in responses if r.maturity >= 3)
        non_compliant = sum(1 for r in responses if r.maturity < 3 and r.maturity > 0)
        in_progress = sum(1 for r in responses if r.maturity <= 1)
        compliance_rate = (compliant / total * 100) if total > 0 else 0

        ws_summary["A8"] = "Métrica"
        ws_summary["B8"] = "Valor"
        ws_summary["C8"] = "Porcentaje"

        ws_summary["A9"] = "Conformes"
        ws_summary["B9"] = compliant
        ws_summary["C9"] = f"={B9}/{B11}*100"

        ws_summary["A10"] = "No Conformes"
        ws_summary["B10"] = non_compliant

        ws_summary["A11"] = "En Progreso"
        ws_summary["B11"] = in_progress

        ws_summary["A12"] = "Total"
        ws_summary["B12"] = total
        ws_summary["C12"] = compliance_rate

        # === HOJA 2: DETALLE POR CONTROL ===
        ws_detail = wb.create_sheet(title="Detalle por Control")

        # Headers
        headers = ["Control", "Dominio", "Madurez", "Estado", "Notas"]
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border

        # Datos
        for row, response in enumerate(responses, 2):
            ctrl = (
                session.get(ControlDefinition, response.control_definition_id)
                if response.control_definition_id
                else None
            )
            ws_detail.cell(row=row, column=1, value=ctrl.code if ctrl else "N/A")
            ws_detail.cell(row=row, column=2, value=ctrl.domain if ctrl else "N/A")
            ws_detail.cell(row=row, column=3, value=response.maturity)

            status = (
                "No Conforme"
                if response.maturity < 3
                else "Conforme"
                if response.maturity >= 3
                else "N/A"
            )
            ws_detail.cell(row=row, column=4, value=status)
            ws_detail.cell(row=row, column=5, value=response.notes or "")

        # Ajustar columnas
        for col in ws_detail.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_detail.column_dimensions[column].width = adjusted_width

        # === HOJA 3: GRÁFICOS ===
        ws_chart = wb.create_sheet(title="Gráficos")

        # Gráfico de barras
        chart = BarChart()
        chart.title = "Distribución de Estados"
        chart.type = "col"

        ws_chart["A1"] = "Estado"
        ws_chart["B1"] = "Cantidad"
        ws_chart["A2"] = "Conformes"
        ws_chart["B2"] = compliant
        ws_chart["A3"] = "No Conformes"
        ws_chart["B3"] = non_compliant
        ws_chart["A4"] = "En Progreso"
        ws_chart["B4"] = in_progress

        data = Reference(ws_chart, min_col=2, min_row=1, max_row=4)
        cats = Reference(ws_chart, min_col=1, min_row=2, max_row=4)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws_chart.add_chart(chart, "E1")

        # Guardar en buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{evaluation.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )


@router.get("/excel-comparativo/{eval1_id}/{eval2_id}")
def export_comparison_excel(eval1_id: str, eval2_id: str, request: Request):
    """Exportar comparativa de dos evaluaciones a Excel"""
    session_id = request.cookies.get("session_id")
    user = get_current_user(session_id)
    if not user:
        raise HTTPException(status_code=401, detail="No autenticado")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    with Session(engine) as session:
        eval1 = session.get(Evaluation, eval1_id)
        eval2 = session.get(Evaluation, eval2_id)

        if not eval1 or not eval2:
            raise HTTPException(status_code=404, detail="Evaluaciones no encontradas")

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativa"

        # Título
        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value = f"COMPARATIVA: {eval1.name} vs {eval2.name}"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")

        # Headers
        ws["A3"] = "Métrica"
        ws["B3"] = eval1.name
        ws["C3"] = eval2.name
        ws["D3"] = "Diferencia"
        ws["E3"] = "Cambio"

        for col in range(1, 6):
            ws.cell(row=3, column=col).font = Font(bold=True)

        # Datos (simulados para demo)
        ws["A4"] = "Score"
        ws["B4"] = 65
        ws["C4"] = 78
        ws["D4"] = "=C4-B4"
        ws["E4"] = "Mejora"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Comparativa_{eval1.name[:10]}_vs_{eval2.name[:10]}.xlsx"

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
